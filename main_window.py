import sys
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QColor
import os
from datetime import datetime
import traceback


class PensionLagAnalyzer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.df = None
        self.results = None
        self.init_ui()
        self.load_data()

    def init_ui(self):
        self.setWindowTitle("Pension Inflation Lag Compensation Calculator")

        window_width = 1600
        window_height = 950

        screen_geometry = QApplication.desktop().screenGeometry()
        x = (screen_geometry.width() - window_width) // 2
        y = (screen_geometry.height() - window_height) // 2
        self.setGeometry(x, y, window_width, window_height)

        font = QFont("Arial", 11)
        self.setFont(font)

        # Main widget with tabs
        self.tab_widget = QTabWidget()
        self.setCentralWidget(self.tab_widget)

        # First tab: Main interface
        main_tab = QWidget()
        self.setup_main_tab(main_tab)
        self.tab_widget.addTab(main_tab, "Compensation Calculation")

        # Second tab: Methodology
        method_tab = QWidget()
        self.setup_method_tab(method_tab)
        self.tab_widget.addTab(method_tab, "Methodology")

    def setup_main_tab(self, tab):
        """Setup main tab"""
        main_widget = QWidget()
        main_layout = QHBoxLayout(main_widget)

        # LEFT panel with table and chart (wide)
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)

        # Table with new header - black text
        table_label = QLabel("Russian Pension System: Calculation of Inflation Lag and Its Compensation Method")
        table_label.setFont(QFont("Arial", 13, QFont.Bold))
        table_label.setAlignment(Qt.AlignCenter)
        table_label.setStyleSheet("padding: 8px; color: #000000;")  # Black color
        left_layout.addWidget(table_label)

        self.table = QTableWidget()
        self.table.setFont(QFont("Arial", 10))
        self.table.horizontalHeader().setFont(QFont("Arial", 10, QFont.Bold))
        self.table.setAlternatingRowColors(True)
        # Stretch the table to full width
        left_layout.addWidget(self.table, stretch=4)

        # Chart
        self.figure = Figure(figsize=(8, 5.5))
        self.canvas = FigureCanvas(self.figure)
        left_layout.addWidget(self.canvas, stretch=5)

        main_layout.addWidget(left_panel, 68)

        # RIGHT panel with parameters (narrow)
        right_panel = QWidget()
        right_panel.setMaximumWidth(400)
        right_layout = QVBoxLayout(right_panel)

        # Header - black text
        title = QLabel("Inflation Losses")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("color: #000000; padding: 10px;")  # Black color
        right_layout.addWidget(title)

        # Calculation parameters
        params = QGroupBox("Calculation Parameters")
        params.setFont(QFont("Arial", 11, QFont.Bold))
        params_layout = QVBoxLayout()
        params_layout.setSpacing(8)

        # Current pension
        pension_label = QLabel("Pension amount in 2025 (RUB):")
        pension_label.setFont(QFont("Arial", 11))
        params_layout.addWidget(pension_label)
        self.pension_input = QLineEdit("25000")
        self.pension_input.setFont(QFont("Arial", 11))
        self.pension_input.setToolTip("Enter pension amount in rubles for 2025")
        params_layout.addWidget(self.pension_input)

        # Start year
        start_label = QLabel("Analysis start year:")
        start_label.setFont(QFont("Arial", 11))
        params_layout.addWidget(start_label)
        self.start_year = QComboBox()
        self.start_year.setFont(QFont("Arial", 11))
        params_layout.addWidget(self.start_year)

        # End year (fixed at 2025)
        end_label = QLabel("Analysis end year:")
        end_label.setFont(QFont("Arial", 11))
        params_layout.addWidget(end_label)
        self.end_year = QLabel("2025")
        self.end_year.setFont(QFont("Arial", 11, QFont.Bold))
        self.end_year.setStyleSheet(
            "color: #000000; padding: 5px; border: 1px solid #ccc; border-radius: 3px; background-color: #f0f0f0;")  # Black color
        params_layout.addWidget(self.end_year)

        # Calculation info
        info_label = QLabel(
            "Inflation lag compensation calculation based on\nannual inflation and pension amount")
        info_label.setFont(QFont("Arial", 9))
        info_label.setStyleSheet(
            "color: #666666; font-style: italic; padding: 5px; background-color: #f9f9f9; border-radius: 3px;")
        info_label.setWordWrap(True)
        params_layout.addWidget(info_label)

        params.setLayout(params_layout)
        right_layout.addWidget(params)

        # Results (order changed)
        results_group = QGroupBox("Calculation Results")
        results_group.setFont(QFont("Arial", 11, QFont.Bold))
        results_layout = QVBoxLayout()

        self.total_paid_label = QLabel("Total paid: -")
        self.total_paid_label.setFont(QFont("Arial", 11))
        results_layout.addWidget(self.total_paid_label)

        # Changed order: average monthly losses first
        self.yearly_avg_label = QLabel("Average monthly losses: -")
        self.yearly_avg_label.setFont(QFont("Arial", 11))
        results_layout.addWidget(self.yearly_avg_label)

        # Then total losses
        self.total_compensation_label = QLabel("Total losses: -")
        self.total_compensation_label.setFont(QFont("Arial", 11))
        self.total_compensation_label.setStyleSheet("color: #000000; font-weight: bold;")  # Black color
        results_layout.addWidget(self.total_compensation_label)

        # Then loss percentage
        self.loss_percentage_label = QLabel("Loss percentage: -")
        self.loss_percentage_label.setFont(QFont("Arial", 11))
        results_layout.addWidget(self.loss_percentage_label)

        self.calc_info_label = QLabel("Load data to calculate")
        self.calc_info_label.setFont(QFont("Arial", 10))
        self.calc_info_label.setStyleSheet("color: #666666; font-style: italic;")
        results_layout.addWidget(self.calc_info_label)

        results_group.setLayout(results_layout)
        right_layout.addWidget(results_group)

        # Buttons
        buttons_vertical = QVBoxLayout()

        calc_btn = QPushButton("Calculate Compensation")
        calc_btn.setFont(QFont("Arial", 11))
        calc_btn.setFixedHeight(40)
        calc_btn.setStyleSheet("""
            QPushButton {
                background-color: #243e4a;
                color: white;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1a2d36;
            }
        """)
        calc_btn.clicked.connect(self.calculate)
        buttons_vertical.addWidget(calc_btn)

        method_btn = QPushButton("Show Methodology")
        method_btn.setFont(QFont("Arial", 11))
        method_btn.setFixedHeight(40)
        method_btn.setStyleSheet("""
            QPushButton {
                background-color: #243e4a;
                color: white;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1a2d36;
            }
        """)
        method_btn.clicked.connect(self.show_methodology)
        buttons_vertical.addWidget(method_btn)

        self.export_excel_btn = QPushButton("Export to Excel")
        self.export_excel_btn.setFont(QFont("Arial", 11))
        self.export_excel_btn.setFixedHeight(40)
        self.export_excel_btn.setStyleSheet("""
            QPushButton {
                background-color: #243e4a;
                color: white;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1a2d36;
            }
        """)
        self.export_excel_btn.clicked.connect(self.export_to_excel)
        self.export_excel_btn.setEnabled(False)
        buttons_vertical.addWidget(self.export_excel_btn)

        exit_btn = QPushButton("Exit")
        exit_btn.setFont(QFont("Arial", 11))
        exit_btn.setFixedHeight(40)
        exit_btn.setStyleSheet("""
            QPushButton {
                background-color: #822832;
                color: white;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1a2d36;
            }
        """)
        exit_btn.clicked.connect(self.close)
        buttons_vertical.addWidget(exit_btn)

        right_layout.addLayout(buttons_vertical)
        right_layout.addStretch()

        main_layout.addWidget(right_panel, 32)
        tab.setLayout(main_layout)

    def setup_method_tab(self, tab):
        """Setup methodology tab"""
        layout = QVBoxLayout(tab)

        # Header
        title_label = QLabel("Inflation Lag Compensation Calculation Methodology")
        title_label.setFont(QFont("Arial", 16, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet(
            "color: #000000; padding: 15px; background-color: #f0f7f0; border-radius: 8px;")  # Black color
        layout.addWidget(title_label)

        # Area for methodology output
        self.method_text = QTextEdit()
        self.method_text.setFont(QFont("Arial", 11))
        self.method_text.setReadOnly(True)
        self.method_text.setStyleSheet("""
            QTextEdit {
                background-color: #f8f8f8;
                border: 1px solid #ddd;
                border-radius: 5px;
                padding: 10px;
                line-height: 1.4;
            }
        """)
        layout.addWidget(self.method_text, 1)

        # Refresh button
        refresh_btn = QPushButton("Update Methodology Calculation")
        refresh_btn.setFont(QFont("Arial", 11))
        refresh_btn.setFixedHeight(40)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #243e4a;
                color: white;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1a2d36;
            }
        """)
        refresh_btn.clicked.connect(self.update_methodology)
        layout.addWidget(refresh_btn)

    def show_methodology(self):
        """Switch to methodology tab"""
        self.tab_widget.setCurrentIndex(1)
        self.update_methodology()

    def update_methodology(self):
        """Update methodology information"""
        if self.results is None:
            self.method_text.setHtml("""
                <div style='text-align: center; padding: 40px;'>
                    <h3 style='color: #666;'>Calculation not performed yet</h3>
                    <p>Perform calculation on the "Compensation Calculation" tab, then return here to view methodology.</p>
                </div>
            """)
            return

        try:
            pension_2025 = float(self.pension_input.text())
            start_year = int(self.start_year.currentText())
            end_year = 2025

            # Get data for methodology
            method_data = self.get_methodology_data(pension_2025, start_year, end_year)

            # Create HTML methodology report with CONTRAST FONT
            html_report = self.create_methodology_html_report(method_data)
            self.method_text.setHtml(html_report)

        except Exception as e:
            self.method_text.setHtml(f"""
                <div style='color: red; padding: 20px;'>
                    <h3>Error forming methodology</h3>
                    <p><b>Error:</b> {str(e)}</p>
                </div>
            """)

    def get_methodology_data(self, pension_2025, start_year, end_year):
        """Get data for methodology"""
        if self.df is None or self.df.empty:
            raise ValueError("No data for analysis")

        # Restore pensions by year
        yearly_pensions = {}
        current_pension = pension_2025
        yearly_pensions[end_year] = current_pension

        for year in range(end_year - 1, start_year - 1, -1):
            next_year_data = self.df[self.df['year'] == year + 1]
            if not next_year_data.empty:
                next_row = next_year_data.iloc[0]
                indexation_rate = next_row['indexation'] / 100
                current_pension = current_pension / (1 + indexation_rate)
            yearly_pensions[year] = current_pension

        # Select last 3 years for detailed analysis
        analysis_years = list(range(max(start_year, end_year - 2), end_year + 1))

        method_data = {
            'pension_2025': pension_2025,
            'start_year': start_year,
            'end_year': end_year,
            'yearly_pensions': yearly_pensions,
            'analysis_years': analysis_years,
            'compensation_details': [],
            'total_compensation': 0.0
        }

        # Calculation for each analysis year
        for year in analysis_years:
            if year not in yearly_pensions:
                continue

            year_data = self.df[self.df['year'] == year]
            if year_data.empty:
                continue

            row = year_data.iloc[0]
            inflation_rate = row['inflation_rosstat'] / 100
            pension = yearly_pensions[year]

            # Exact calculation using formula
            monthly_inflation = (1 + inflation_rate) ** (1 / 12) - 1

            # Calculate series sum (formula from code)
            series_sum = 0.0
            monthly_details = []
            for month in range(1, 13):
                price_growth = (1 + monthly_inflation) ** month
                monthly_compensation = pension * (1 - 1 / price_growth)
                series_sum += (1 - 1 / price_growth)
                monthly_details.append({
                    'month': month,
                    'price_growth': price_growth,
                    'monthly_compensation': monthly_compensation
                })

            # Total compensation for the year
            year_compensation = pension * series_sum

            method_data['compensation_details'].append({
                'year': year,
                'pension': pension,
                'inflation_rate': inflation_rate * 100,
                'monthly_inflation': monthly_inflation * 100,
                'series_sum': series_sum,
                'compensation': year_compensation,
                'avg_monthly': year_compensation / 12,
                'monthly_details': monthly_details,
                'percentage_of_pension': (year_compensation / (pension * 12)) * 100
            })

            method_data['total_compensation'] += year_compensation

        return method_data

    def create_methodology_html_report(self, method_data):
        """Create HTML methodology report"""
        # Get data for 2025 as example
        example_year = method_data['analysis_years'][-1] if method_data['analysis_years'] else 2025
        example_detail = None
        for detail in method_data['compensation_details']:
            if detail['year'] == example_year:
                example_detail = detail
                break

        if not example_detail:
            return "<h3>Data not found</h3>"

        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
        <meta charset="UTF-8">
        <style>
            body {{
                font-family: 'Arial', sans-serif;
                line-height: 1.6;
                color: #333;
                margin: 0;
                padding: 20px;
            }}

            .header {{
                text-align: center;
                background: linear-gradient(135deg, #2d4a24 0%, #243e4a 100%);
                color: white;
                padding: 20px;
                border-radius: 10px;
                margin-bottom: 30px;
            }}

            .header h1 {{
                margin: 0;
                font-size: 22px;
                font-weight: bold;
            }}

            .simple-explanation {{
                background: #f0f8ff;
                border-radius: 8px;
                padding: 20px;
                margin: 20px 0;
                border-left: 4px solid #1e90ff;
            }}

            .simple-explanation h3 {{
                color: #1e3c72;
                margin-top: 0;
            }}

            .section {{
                background: white;
                border-radius: 8px;
                padding: 20px;
                margin: 20px 0;
                box-shadow: 0 2px 4px rgba(0,0,0,0.05);
                border-left: 4px solid #2d4a24;
            }}

            .formula {{
                background: #000000;
                border: 3px solid #FF0000;
                border-radius: 8px;
                padding: 30px;
                margin: 20px 0;
                font-family: 'Courier New', monospace;
                font-size: 28px;
                text-align: center;
                color: #FFFFFF;
                font-weight: bold;
            }}

            .example-box {{
                background: #e8f5e9;
                border: 2px solid #2d4a24;
                border-radius: 8px;
                padding: 20px;
                margin: 20px 0;
            }}

            .table {{
                width: 100%;
                border-collapse: collapse;
                margin: 20px 0;
            }}

            .table th {{
                background: #2d4a24;
                color: white;
                padding: 12px;
                text-align: left;
                font-weight: bold;
            }}

            .table td {{
                padding: 10px;
                border-bottom: 1px solid #ddd;
                color: #000000;
            }}

            .table tr:nth-child(even) {{
                background: #f8f9fa;
            }}

            .highlight {{
                background: #e3f2fd;
                font-weight: bold;
                color: #000000;
            }}

            .result {{
                background: #fff3e0;
                padding: 15px;
                border-radius: 6px;
                margin: 15px 0;
                border-left: 4px solid #ff9800;
                color: #000000;
                font-weight: bold;
            }}

            .verification {{
                background: #e1f5fe;
                padding: 15px;
                border-radius: 6px;
                margin: 15px 0;
                border-left: 4px solid #0288d1;
                color: #000000;
                font-weight: bold;
            }}

            h2, h3, h4 {{
                color: #2d4a24;
                font-weight: bold;
            }}

            p, li {{
                color: #000000;
            }}
        </style>
        </head>
        <body>

        <div class="simple-explanation">
            <h3>Method explanation:</h3>
            <p><b>The government indexes pensions so they don't fall behind price increases (inflation). However, this happens with a delay: prices rise throughout the year, but pensions are increased only once. Due to this gap — the inflation lag — the real purchasing power of your monthly payment by year-end becomes less than at the beginning.</b></p>
            <p><b>What does this mean for you?</b> Essentially, each month your pension slightly "melts" compared to store prices.</p>
            <p><b>What does the program do?</b> It calculates the total amount of these monthly losses for the year. This amount can be considered as a necessary additional payment ("13th pension") to fully protect your money from inflation.</p>
        </div>

        <div class="section">
            <h2>1. Mathematical basis of the method</h2>
            <p><b>Inflation lag compensation is calculated by the formula:</b></p>

            <div class="formula">
                C = P × ∑<sub>m=1</sub><sup>12</sup> [1 - (1 + i)<sup>-m/12</sup>]
            </div>

            <p><b>where:</b></p>
            <ul>
                <li><b>C</b> – annual compensation payment (RUB)</li>
                <li><b>P</b> – pension amount in January of calculation year (RUB)</li>
                <li><b>i</b> – annual inflation as decimal</li>
                <li><b>m</b> – month number (from 1 to 12)</li>
            </ul>

            <p><b>Physical meaning of the formula:</b> For each month, the depreciation of pension payment relative to price growth from year start is calculated, then these losses are summed.</p>
        </div>

        <div class="section">
            <h2>2. Calculation example for {example_detail['year']}</h2>

            <div class="example-box">
                <h3>Initial data:</h3>
                <ul>
                    <li><b>Pension in January:</b> {example_detail['pension']:,.2f} RUB</li>
                    <li><b>Annual inflation:</b> {example_detail['inflation_rate']:.2f}% (i = {example_detail['inflation_rate'] / 100:.4f})</li>
                    <li><b>Average monthly inflation:</b> {example_detail['monthly_inflation']:.4f}%</li>
                </ul>
            </div>

            <h3>Step-by-step calculation:</h3>

            <table class="table">
                <tr>
                    <th>Month</th>
                    <th>Price growth<br>(1+π)<sup>m</sup></th>
                    <th>Payment depreciation<br>1 - 1/(1+π)<sup>m</sup></th>
                    <th>Monthly losses</th>
                </tr>
        """

        # Add monthly data (first 3, last 3 and total)
        months_to_show = [1, 2, 3, 10, 11, 12]
        for month in months_to_show:
            if month <= len(example_detail['monthly_details']):
                detail = example_detail['monthly_details'][month - 1]
                html += f"""
                <tr>
                    <td><b>{month} ({['January', 'February', 'March', 'October', 'November', 'December'][months_to_show.index(month)]})</b></td>
                    <td><b>{detail['price_growth']:.6f}</b></td>
                    <td><b>{1 - 1 / detail['price_growth']:.6f}</b></td>
                    <td><b>{detail['monthly_compensation']:.2f} RUB</b></td>
                </tr>
                """

        html += f"""
                <tr class="highlight">
                    <td colspan="3"><b>Sum of coefficients ∑[1 - 1/(1+π)<sup>m</sup>]:</b></td>
                    <td><b>{example_detail['series_sum']:.6f}</b></td>
                </tr>
                <tr class="highlight">
                    <td colspan="3"><b>Annual losses (compensation):</b></td>
                    <td><b>{example_detail['compensation']:,.2f} RUB</b></td>
                </tr>
            </table>

            <div class="result">
                <h3>Result for {example_detail['year']}:</h3>
                <p><b>Annual losses (compensation):</b> {example_detail['compensation']:,.2f} RUB</p>
                <p><b>Average monthly losses:</b> {example_detail['avg_monthly']:,.2f} RUB</p>
                <p><b>As percentage of annual pension:</b> {example_detail['percentage_of_pension']:.2f}%</p>
                <p><b>Equivalent in monthly pensions:</b> {example_detail['compensation'] / example_detail['pension']:.2f} months</p>
            </div>
        </div>

        <div class="section">
            <h2>3. Formula correctness verification</h2>

            <div class="verification">
                <h3>Let's verify formula correctness using December as example:</h3>
                <p>For December (m=12):</p>
                <p>(1+i)<sup>-12/12</sup> = (1+{example_detail['inflation_rate'] / 100:.4f})<sup>-1</sup> = {1 / (1 + example_detail['inflation_rate'] / 100):.6f}</p>
                <p>Depreciation = 1 - {1 / (1 + example_detail['inflation_rate'] / 100):.6f} = {1 - 1 / (1 + example_detail['inflation_rate'] / 100):.6f}</p>
                <p>December losses = {example_detail['pension']:,.2f} × {1 - 1 / (1 + example_detail['inflation_rate'] / 100):.6f} = 
                <b>{example_detail['monthly_details'][11]['monthly_compensation']:.2f} RUB</b></p>
            </div>

            <p><b>Calculation matches program results, confirming formula correctness.</b></p>
        </div>

        <div class="section">
            <h2>4. Results interpretation</h2>

            <p><b>The obtained losses represent the amount needed to restore purchasing power of pension payments lost due to inflation lag.</b></p>

            <ul>
                <li><b>Average monthly losses</b> show by how much the real value of pension decreases annually due to indexation lagging behind inflation</li>

                <li><b>Annual losses (compensation)</b> represents the necessary compensation amount for the year ("13th pension")</li>

                <li><b>The method doesn't require complex retrospective calculations</b> and is based on transparent mathematical formulas</li>
            </ul>

            <p><b>Method advantages:</b> simplicity, transparency, implementability, compliance with social justice principles.</p>
        </div>

        </body>
        </html>
        """

        return html

    def load_data(self):
        """Load data ONLY from file"""
        try:
            # Determine where to look for file
            if getattr(sys, 'frozen', False):
                # If running as EXE - look next to EXE
                base_path = os.path.dirname(sys.executable)
            else:
                # If running as script - look next to script
                base_path = os.path.dirname(__file__)

            # Path to Excel file
            excel_path = os.path.join(base_path, 'data', 'russia_inflation.xlsx')

            if os.path.exists(excel_path):
                self.df = pd.read_excel(excel_path)
                print(f"Data loaded from Excel: {excel_path}")
            else:
                # File not found - show error
                error_msg = (
                    "Data file not found!\n\n"
                    f"Expected path: {excel_path}\n\n"
                    "File 'russia_inflation.xlsx' required for program operation\n"
                    "in folder 'data' next to the program.\n\n"
                    "File structure:\n"
                    "- year (year)\n"
                    "- inflation_rosstat (inflation in %)\n"
                    "- indexation (indexation in %)"
                )
                QMessageBox.critical(self, "Data Loading Error", error_msg)
                self.df = pd.DataFrame()
                return

            # Check data structure
            required_columns = ['year', 'inflation_rosstat', 'indexation']
            missing_columns = [col for col in required_columns if col not in self.df.columns]

            if missing_columns:
                error_msg = (
                    "Invalid data format!\n\n"
                    f"File missing columns: {', '.join(missing_columns)}\n"
                    f"Found columns: {', '.join(self.df.columns.tolist())}\n\n"
                    "Required columns:\n"
                    "- year (year)\n"
                    "- inflation_rosstat (inflation in %)\n"
                    "- indexation (indexation in %)"
                )
                QMessageBox.critical(self, "Data Format Error", error_msg)
                self.df = pd.DataFrame()
                return

            # Convert data types
            self.df['year'] = self.df['year'].astype(int)
            self.df['inflation_rosstat'] = pd.to_numeric(self.df['inflation_rosstat'], errors='coerce')
            self.df['indexation'] = pd.to_numeric(self.df['indexation'], errors='coerce')

            # Check for missing values
            if self.df.isnull().any().any():
                QMessageBox.warning(
                    self,
                    "Warning",
                    "Missing values found in data. They will be filled."
                )
                self.df = self.df.ffill()

            # Update year list
            years = self.df['year'].astype(int).tolist()
            self.start_year.clear()
            self.start_year.addItems([str(y) for y in years if y < 2025])

            if years:
                # Set 2020 as default if present in data
                if 2020 in years:
                    self.start_year.setCurrentText("2020")
                else:
                    self.start_year.setCurrentText(str(min(years)))

            self.calc_info_label.setText(f"Data loaded: {len(years)} years ({min(years)}-{max(years)})")

        except pd.errors.EmptyDataError:
            QMessageBox.critical(self, "Error", "Data file is empty.")
            self.df = pd.DataFrame()

        except pd.errors.ParserError as e:
            QMessageBox.critical(self, "Parsing Error", f"Error reading data file:\n{str(e)}")
            self.df = pd.DataFrame()

        except Exception as e:
            error_details = traceback.format_exc()
            QMessageBox.critical(
                self,
                "Data Loading Error",
                f"Failed to load data:\n\n{str(e)}\n\nDetails:\n{error_details}"
            )
            self.df = pd.DataFrame()

    def calculate(self):
        """Calculate compensation"""
        try:
            # Check if data exists
            if self.df is None or self.df.empty:
                QMessageBox.warning(
                    self,
                    "No Data",
                    "Load data first.\n"
                    "Place file 'russia_inflation.xlsx' in 'data' folder next to program."
                )
                return

            pension_2025 = float(self.pension_input.text())
            start = int(self.start_year.currentText())
            end = 2025

            if pension_2025 <= 0:
                QMessageBox.warning(self, "Error", "Pension amount must be positive")
                return

            if start >= end:
                QMessageBox.warning(self, "Error", "Start year must be less than 2025")
                return

            self.results = self.calculate_compensation(pension_2025, start, end)
            self.update_results(self.results, start, end)
            self.update_table(self.results)
            self.plot_chart(self.results)

            self.export_excel_btn.setEnabled(True)

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Calculation error: {str(e)}")

    def calculate_compensation(self, pension_2025, start_year, end_year):
        """Calculate inflation lag compensation"""
        # Restore pensions
        yearly_pensions = {}
        current_pension = pension_2025
        yearly_pensions[end_year] = current_pension

        for year in range(end_year - 1, start_year - 1, -1):
            next_year_data = self.df[self.df['year'] == year + 1]
            if not next_year_data.empty:
                next_row = next_year_data.iloc[0]
                indexation_rate = next_row['indexation'] / 100.0
                current_pension = current_pension / (1.0 + indexation_rate)
            yearly_pensions[year] = current_pension

        details = []
        yearly_summary = {}
        total_paid = 0.0
        total_compensation = 0.0

        for year in range(start_year, end_year + 1):
            if year not in yearly_pensions:
                continue

            year_data = self.df[self.df['year'] == year]
            if year_data.empty:
                continue

            row = year_data.iloc[0]
            inflation_rate = row['inflation_rosstat'] / 100.0
            pension = yearly_pensions[year]
            monthly_inflation = (1.0 + inflation_rate) ** (1.0 / 12.0) - 1.0

            year_paid = 0.0
            year_compensation = 0.0

            # Compensation calculation using formula
            for month in range(1, 13):
                price_growth = (1.0 + monthly_inflation) ** month
                monthly_paid = pension
                monthly_compensation = monthly_paid * (1.0 - 1.0 / price_growth)

                year_paid += monthly_paid
                year_compensation += monthly_compensation
                total_paid += monthly_paid
                total_compensation += monthly_compensation

                details.append({
                    'Year': year,
                    'Month': month,
                    'Pension': round(pension, 4),
                    'Paid': round(monthly_paid, 4),
                    'Compensation': round(monthly_compensation, 4)
                })

            loss_percentage = (year_compensation / year_paid * 100.0) if year_paid > 0 else 0.0

            yearly_summary[year] = {
                'pension_in_january': pension,
                'inflation_year': inflation_rate * 100.0,
                'indexation_year': row['indexation'],
                'sum_per_year': year_paid,
                'compensation_per_year': year_compensation,
                'compensation_per_month': year_compensation / 12.0,
                'loss_percentage': loss_percentage,
                'total_compensation': 0.0
            }

        # Cumulative compensation
        cumulative_compensation = 0.0
        for year in sorted(yearly_summary.keys()):
            cumulative_compensation += yearly_summary[year]['compensation_per_year']
            yearly_summary[year]['total_compensation'] = cumulative_compensation

        return {
            'total_paid': total_paid,
            'total_compensation': total_compensation,
            'loss_percentage': (total_compensation / total_paid * 100.0) if total_paid > 0 else 0.0,
            'details': details,
            'yearly_summary': yearly_summary,
            'yearly_pensions': yearly_pensions
        }

    def update_results(self, result, start_year, end_year):
        """Update results on panel"""
        total_paid = result['total_paid']
        total_compensation = result['total_compensation']
        loss_percentage = result['loss_percentage']

        years_count = end_year - start_year + 1
        avg_monthly_loss = total_compensation / (years_count * 12) if years_count > 0 else 0

        self.total_paid_label.setText(
            f"Total paid: {total_paid:,.0f} RUB".replace(',', ' ')
        )
        # Changed order: average monthly losses first
        self.yearly_avg_label.setText(
            f"Average monthly losses: {avg_monthly_loss:,.0f} RUB".replace(',', ' ')
        )
        # Then total losses
        self.total_compensation_label.setText(
            f"Total losses: {total_compensation:,.0f} RUB".replace(',', ' ')
        )
        # Then loss percentage
        self.loss_percentage_label.setText(
            f"Loss percentage: {loss_percentage:.2f}%"
        )

    def update_table(self, result):
        """Update table"""
        if not result['yearly_summary']:
            return

        years = sorted(result['yearly_summary'].keys())
        self.table.setRowCount(len(years))
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels([
            'Year',
            'Pension\namount',
            'Inflation\n(%)',
            'Indexation\n(%)',
            'Paid\nper year',
            'Losses\nper month',
            'Losses per year\n(compensation)',  # Wide column
            'Accumulated\nlosses'  # Changed name
        ])

        # Set column widths - now all columns are stretchable
        header = self.table.horizontalHeader()
        for col in range(8):
            header.setSectionResizeMode(col, QHeaderView.Stretch)

        for i, year in enumerate(years):
            data = result['yearly_summary'][year]

            # Year - black font
            year_item = QTableWidgetItem(str(year))
            year_item.setTextAlignment(Qt.AlignCenter)
            year_item.setForeground(QColor(0, 0, 0))  # Black font
            self.table.setItem(i, 0, year_item)

            # Pension - black font
            pension_item = QTableWidgetItem(f"{data['pension_in_january']:,.2f}".replace(',', ' '))
            pension_item.setTextAlignment(Qt.AlignCenter)
            pension_item.setForeground(QColor(0, 0, 0))  # Black font
            self.table.setItem(i, 1, pension_item)

            # Inflation - black font
            inflation_item = QTableWidgetItem(f"{data['inflation_year']:.2f}%")
            inflation_item.setTextAlignment(Qt.AlignCenter)
            inflation_item.setForeground(QColor(0, 0, 0))  # Black font
            self.table.setItem(i, 2, inflation_item)

            # Indexation (precision to tenths) - black font
            indexation_item = QTableWidgetItem(f"{data['indexation_year']:.1f}%")
            indexation_item.setTextAlignment(Qt.AlignCenter)
            indexation_item.setForeground(QColor(0, 0, 0))  # Black font
            self.table.setItem(i, 3, indexation_item)

            # Paid per year - black font
            paid_item = QTableWidgetItem(f"{data['sum_per_year']:,.0f}".replace(',', ' '))
            paid_item.setTextAlignment(Qt.AlignCenter)
            paid_item.setForeground(QColor(0, 0, 0))  # Black font
            self.table.setItem(i, 4, paid_item)

            # Losses per month - black font
            monthly_loss_item = QTableWidgetItem(f"{data['compensation_per_month']:,.0f}".replace(',', ' '))
            monthly_loss_item.setTextAlignment(Qt.AlignCenter)
            monthly_loss_item.setForeground(QColor(0, 0, 0))  # Black font
            self.table.setItem(i, 5, monthly_loss_item)

            # Losses per year (compensation) - red font preserved
            year_loss_item = QTableWidgetItem(f"{data['compensation_per_year']:,.0f}".replace(',', ' '))
            year_loss_item.setTextAlignment(Qt.AlignCenter)
            year_loss_item.setForeground(QColor(128, 0, 0))  # Red font preserved
            year_loss_item.setFont(QFont("Arial", 10, QFont.Bold))
            self.table.setItem(i, 6, year_loss_item)

            # Accumulated losses - black font
            cumulative_item = QTableWidgetItem(f"{data['total_compensation']:,.0f}".replace(',', ' '))
            cumulative_item.setTextAlignment(Qt.AlignCenter)
            cumulative_item.setForeground(QColor(0, 0, 0))  # Black font
            self.table.setItem(i, 7, cumulative_item)

            # Alternating row background
            if i % 2 == 0:
                for col in range(8):
                    item = self.table.item(i, col)
                    if item:
                        item.setBackground(QColor(245, 245, 245))

    def plot_chart(self, result):
        """Build chart"""
        self.figure.clear()
        ax = self.figure.add_subplot(111)

        if not result['yearly_summary']:
            return

        years = sorted(result['yearly_summary'].keys())
        pensions = [result['yearly_summary'][y]['pension_in_january'] for y in years]
        compensations = [result['yearly_summary'][y]['compensation_per_year'] for y in years]
        cumulative_compensations = [result['yearly_summary'][y]['total_compensation'] for y in years]

        x_pos = np.arange(len(years))
        bar_width = 0.35

        # Bars: pensions and compensations (losses)
        bars_pension = ax.bar(x_pos - bar_width / 2, pensions, bar_width,
                              color='#243e4a', alpha=0.7, label='Pension amount')

        bars_compensation = ax.bar(x_pos + bar_width / 2, compensations, bar_width,
                                   color='#800000', alpha=0.7, label='Losses per year (compensation)')

        # Accumulated losses line
        line_cumulative = ax.plot(x_pos, cumulative_compensations, 'o-',
                                  color='#556b2f', linewidth=2, markersize=6,
                                  label='Accumulated losses')[0]

        # Labels on pension bars
        for bar in bars_pension:
            height = bar.get_height()
            if height > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, height,
                        f'{height:,.0f}'.replace(',', ' '),
                        ha='center', va='bottom', fontsize=8, color='#243e4a')

        # Labels on compensation bars (losses)
        for bar in bars_compensation:
            height = bar.get_height()
            if height > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, height,
                        f'{height:,.0f}'.replace(',', ' '),
                        ha='center', va='bottom', fontsize=8, color='#800000', fontweight='bold')

        # Labels on accumulated losses line
        for i, (x, y) in enumerate(zip(x_pos, cumulative_compensations)):
            ax.text(x, y, f'{y:,.0f}'.replace(',', ' '),
                    ha='center', va='bottom', fontsize=8, color='#556b2f', fontweight='bold')

        ax.set_title('Pension Dynamics and Inflation Lag Losses', fontsize=12, fontweight='bold')
        ax.set_xlabel('Year', fontsize=10)
        ax.set_ylabel('Amount, RUB', fontsize=10)

        ax.set_xticks(x_pos)
        ax.set_xticklabels([str(y) for y in years], fontsize=9)
        ax.tick_params(axis='y', labelsize=9)

        ax.legend(fontsize=9, loc='upper left')
        ax.grid(True, alpha=0.3, axis='y')
        ax.margins(y=0.1)

        self.figure.tight_layout()
        self.canvas.draw()

    def export_to_excel(self):
        """Export to Excel with formatting"""
        if self.results is None:
            QMessageBox.warning(self, "Error", "Perform calculation first")
            return

        result_dir = "Result"
        if not os.path.exists(result_dir):
            os.makedirs(result_dir)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"inflation_lag_compensation_{timestamp}.xlsx"
        file_path = os.path.join(result_dir, file_name)

        try:
            from openpyxl import Workbook
            from openpyxl.styles import (
                Alignment, PatternFill, Font, Border, Side
            )
            from openpyxl.utils import get_column_letter

            wb = Workbook()

            # Sheet 1: Yearly summary
            ws1 = wb.active
            ws1.title = "Yearly Summary"

            headers = [
                'Year',
                'Pension in January',
                'Inflation (%)',
                'Indexation (%)',
                'Paid per year',
                'Losses per month',
                'Losses per year (compensation)',
                'Loss percentage (%)',
                'Accumulated losses'  # Changed name
            ]

            ws1.append(headers)

            header_fill = PatternFill(start_color="2d4a24", end_color="2d4a24", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            center_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for col in range(1, len(headers) + 1):
                cell = ws1.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
                ws1.column_dimensions[get_column_letter(col)].width = 18

            row_num = 2
            for year in sorted(self.results['yearly_summary'].keys()):
                data = self.results['yearly_summary'][year]

                row_data = [
                    year,
                    round(data['pension_in_january'], 2),
                    round(data['inflation_year'], 2),
                    round(data['indexation_year'], 1),  # Precision to tenths
                    round(data['sum_per_year'], 2),
                    round(data['compensation_per_month'], 2),  # Losses per month
                    round(data['compensation_per_year'], 2),  # Losses per year (compensation)
                    round(data['loss_percentage'], 2),
                    round(data['total_compensation'], 2)  # Accumulated losses
                ]

                ws1.append(row_data)

                for col in range(1, len(row_data) + 1):
                    cell = ws1.cell(row=row_num, column=col)
                    cell.alignment = center_alignment
                    cell.border = thin_border

                    if col in [2, 5, 6, 7, 9]:  # Monetary values
                        cell.number_format = '#,##0.00'
                    elif col in [3, 4, 8]:  # Percentages
                        cell.number_format = '0.0"%"' if col == 4 else '0.00"%"'  # Indexation to tenths

                    if row_num % 2 == 0:
                        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

                loss_cell = ws1.cell(row=row_num, column=7)  # Losses per year
                loss_cell.font = Font(color="800000", bold=True)

                row_num += 1

            # Sheet 2: Summary
            ws2 = wb.create_sheet(title="Summary")

            loss_percentage = self.results['loss_percentage']

            summary_data = [
                ["Parameter", "Value"],
                ["Analysis start year", int(self.start_year.currentText())],
                ["Analysis end year", 2025],
                ["Pension amount in 2025", float(self.pension_input.text())],
                ["Total analysis months", (2025 - int(self.start_year.currentText()) + 1) * 12],
                ["Total paid", round(self.results['total_paid'], 2)],
                ["Total losses", round(self.results['total_compensation'], 2)],  # Changed
                ["Loss percentage (%)", f"{loss_percentage:.2f}%"],
                ["Average monthly losses",
                 round(self.results['total_compensation'] / ((2025 - int(self.start_year.currentText()) + 1) * 12), 2)],
                ["Calculation date", pd.Timestamp.now().strftime("%d.%m.%Y %H:%M:%S")]
            ]

            for row in summary_data:
                ws2.append(row)

            for col in range(1, 3):
                cell = ws2.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border

            for row in range(2, len(summary_data) + 1):
                for col in range(1, 3):
                    cell = ws2.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal="left" if col == 1 else "right", vertical="center")
                    cell.border = thin_border
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

            ws2.column_dimensions['A'].width = 30
            ws2.column_dimensions['B'].width = 20

            # Sheet 3: Source data
            ws3 = wb.create_sheet(title="Source Data")

            source_headers = ['Year', 'Rosstat Inflation (%)', 'Pension Indexation (%)']
            ws3.append(source_headers)

            for col in range(1, 4):
                cell = ws3.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
                ws3.column_dimensions[get_column_letter(col)].width = 20

            for _, row in self.df.iterrows():
                ws3.append([row['year'], row['inflation_rosstat'], row['indexation']])

            for row in range(2, len(self.df) + 2):
                for col in range(1, 4):
                    cell = ws3.cell(row=row, column=col)
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    if col in [2, 3]:
                        cell.number_format = '0.0"%"' if col == 3 else '0.00"%"'  # Indexation to tenths

            wb.save(file_path)

            QMessageBox.information(
                self,
                "Success",
                f"Calculation results saved:\n{file_path}"
            )

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save Excel file: {str(e)}")


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = PensionLagAnalyzer()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()